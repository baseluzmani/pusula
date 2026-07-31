"""
ETF Holdings Importer
---------------------
Two sources:
1. Historical Excel: data/Funds Database.xlsx
2. Ongoing CSVs/XLS/XLSX: data/etf_holdings_import/input/

File naming for ongoing imports: {PREFIX}_{anything}.csv / .xls / .xlsx
  e.g. AINF_20260610.csv, DFNS_asof_20260609.xlsx

Date is always read from inside the file.
If (etf_fund_id, scraped_date) already exists in DB, file is skipped.
Processed files are moved to data/etf_holdings_import/archive/

Usage:
    python3 scripts/import_etf_holdings.py              # process ongoing files only
    python3 scripts/import_etf_holdings.py --historical  # also import from Excel
"""

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sqlite3
import csv
import re
import xml.etree.ElementTree as ET
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core import config
from core.repo import tickers as _tickers

# ── Config ────────────────────────────────────────────────────────────────────
# Paths come from core.config as absolute Paths; the maps come from the
# instruments table. Nothing here depends on the working directory.

DB_PATH     = str(config.DB_PATH)
IMPORT_DIR  = Path(config.IMPORT_DIR)
ARCHIVE_DIR = Path(config.ARCHIVE_DIR)
EXCEL_PATH  = Path(config.EXCEL_PATH)

FUND_ID_MAP      = _tickers.fund_id_map()
ETF_PROVIDER_MAP = _tickers.etf_provider_map()

# Skip these asset classes in holdings
SKIP_ASSET_CLASSES = {
    'cash and/or derivatives', 'cash', 'futures', 'options',
    'cash collateral and margins',
}

# Skip these tickers — currencies, FX pairs, cash placeholders
SKIP_TICKERS = {
    'AUD', 'CAD', 'CHF', 'CZK', 'EUR', 'GBP', 'HKD', 'JPY', 'KRW',
    'MXN', 'MYR', 'PLN', 'SEK', 'SGD', 'TRY', 'TWD', 'USD', 'ZAR',
    'DKK', 'NOK', 'NZD', 'THB', 'IDR', 'INR', 'BRL', 'MXN', 'ZAR',
    '-', '$', 'CASH', 'XONE',
}

# Sheets to skip in historical Excel (NAV/performance/price data)
SKIP_SHEETS = {
    'Sheet1', 'Sheet2', 'Sheet9', 'AINF', 'DFEU', 'MINE', 'SL Funds',
    'HSBC Pension', 'Burcu Dashboard', 'Ahmet Dashboard',
}


# ── DB Setup ──────────────────────────────────────────────────────────────────

def get_conn():
    return sqlite3.connect(DB_PATH, timeout=30)


def create_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS etf_holdings (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            etf_fund_id     TEXT NOT NULL,
            scraped_date    TEXT NOT NULL,
            ticker          TEXT,
            name            TEXT NOT NULL,
            sector          TEXT,
            asset_class     TEXT,
            weight_pct      REAL,
            market_value    REAL,
            location        TEXT,
            currency        TEXT,
            UNIQUE(etf_fund_id, scraped_date, name)
        );
        CREATE INDEX IF NOT EXISTS idx_etf_holdings_fund_date
            ON etf_holdings(etf_fund_id, scraped_date);
        CREATE INDEX IF NOT EXISTS idx_etf_holdings_ticker
            ON etf_holdings(ticker);
    """)
    conn.commit()


def already_imported(conn, etf_fund_id, scraped_date):
    count = conn.execute(
        "SELECT COUNT(*) FROM etf_holdings WHERE etf_fund_id=? AND scraped_date=?",
        (etf_fund_id, scraped_date)
    ).fetchone()[0]
    return count > 0


def insert_holdings(conn, etf_fund_id, scraped_date, holdings):
    # Ensure isin and sedol columns exist (for ISIN/SEDOL-only providers)
    for col in ('isin TEXT', 'sedol TEXT'):
        try:
            conn.execute(f"ALTER TABLE etf_holdings ADD COLUMN {col}")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # column already exists

    inserted = 0
    errors   = 0
    for h in holdings:
        # Skip currencies, FX pairs and cash placeholders
        ticker = h.get('ticker', '') or ''
        if str(ticker).strip().upper() in SKIP_TICKERS:
            continue
        # Skip cash/derivative asset classes
        asset_class = str(h.get('asset_class', '') or '').lower().strip()
        if asset_class in SKIP_ASSET_CLASSES:
            continue
        try:
            conn.execute("""
                INSERT OR REPLACE INTO etf_holdings
                    (etf_fund_id, scraped_date, ticker, name, sector,
                     asset_class, weight_pct, market_value, location, currency, isin, sedol)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                etf_fund_id, scraped_date,
                h.get('ticker'), h['name'], h.get('sector'),
                h.get('asset_class'), h.get('weight_pct'), h.get('market_value'),
                h.get('location'), h.get('currency'), h.get('isin'), h.get('sedol'),
            ))
            inserted += 1
        except Exception as e:
            errors += 1
            print(f"    Row error ({h.get('name', '?')}): {e}")
    conn.commit()
    return inserted, errors


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_float(val, multiply=1.0):
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('$', '').strip()
        if not s or s in ('-', 'N/A', 'nan', '--'):
            return None
        return float(s) * multiply
    except:
        return None


# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_ishares_csv(filepath):
    """
    iShares CSV format.
    Header rows at top contain date. Data starts after column header row.
    Columns: Ticker, Name, Sector, Asset Class, Market Value, Weight (%), Location, Market Currency
    """
    holdings     = []
    scraped_date = None

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    # Find date in first 5 lines
    for line in lines[:5]:
        m = re.search(r'(\d{2}/\w+/\d{4})', line)
        if m:
            try:
                scraped_date = datetime.strptime(m.group(1), '%d/%b/%Y').strftime('%Y-%m-%d')
                break
            except:
                pass

    # Find header row
    header_idx = None
    for i, line in enumerate(lines):
        if 'Ticker' in line and 'Name' in line and 'Weight' in line:
            header_idx = i
            break

    if header_idx is None:
        return holdings, scraped_date

    reader = csv.DictReader(lines[header_idx:])
    for row in reader:
        ticker    = (row.get('Ticker') or row.get('Issuer Ticker') or '').strip().strip('"')
        name      = (row.get('Name') or '').strip().strip('"')
        sector    = (row.get('Sector') or '').strip().strip('"')
        asset_cls = (row.get('Asset Class') or '').strip().strip('"')
        weight    = safe_float(row.get('Weight (%)'))
        mktval    = safe_float(row.get('Market Value'))
        location  = (row.get('Location') or '').strip().strip('"')
        currency  = (row.get('Market Currency') or '').strip().strip('"')

        if not name or asset_cls.lower() in SKIP_ASSET_CLASSES:
            continue

        holdings.append({
            'ticker': ticker, 'name': name, 'sector': sector,
            'asset_class': asset_cls, 'weight_pct': weight,
            'market_value': mktval, 'location': location, 'currency': currency,
        })

    return holdings, scraped_date


def parse_ishares_xls(filepath):
    """iShares XLS (XML Spreadsheet) format."""
    holdings     = []
    scraped_date = None

    with open(filepath, 'rb') as f:
        content = f.read().decode('utf-8-sig')

    ns   = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    root = ET.fromstring(content)
    rows = root.findall('.//ss:Row', ns)

    def get_vals(row):
        cells = row.findall('ss:Cell', ns)
        return [
            c.find('ss:Data', ns).text or ''
            if c.find('ss:Data', ns) is not None else ''
            for c in cells
        ]

    # Find date
    for row in rows[:8]:
        for v in get_vals(row):
            m = re.search(r'(\d{2}[/-]\w+[/-]\d{4}|\d{4}-\d{2}-\d{2})', str(v or ''))
            if m:
                for fmt in ['%d/%b/%Y', '%d-%b-%Y', '%Y-%m-%d']:
                    try:
                        scraped_date = datetime.strptime(m.group(1), fmt).strftime('%Y-%m-%d')
                        break
                    except:
                        pass
            if scraped_date:
                break
        if scraped_date:
            break

    # Find header row
    headers    = []
    header_idx = None
    for i, row in enumerate(rows):
        vals = get_vals(row)
        if any(v in ('Name', 'Ticker', 'Issuer Ticker') for v in vals):
            headers    = vals
            header_idx = i
            break

    if header_idx is None:
        return holdings, scraped_date

    def col(options):
        for n in options:
            if n in headers:
                return headers.index(n)
        return None

    idx = {
        'ticker':   col(['Ticker', 'Issuer Ticker']),
        'name':     col(['Name']),
        'sector':   col(['Sector']),
        'asset':    col(['Asset Class']),
        'weight':   col(['Weight (%)']),
        'mktval':   col(['Market Value']),
        'location': col(['Location']),
        'currency': col(['Market Currency']),
    }

    for row in rows[header_idx + 1:]:
        vals = get_vals(row)
        if not any(vals):
            continue

        def g(k):
            i = idx[k]
            return vals[i].strip() if i is not None and i < len(vals) else ''

        name      = g('name')
        asset_cls = g('asset')
        if not name or asset_cls.lower() in SKIP_ASSET_CLASSES:
            continue

        holdings.append({
            'ticker':       g('ticker'),
            'name':         name,
            'sector':       g('sector'),
            'asset_class':  asset_cls,
            'weight_pct':   safe_float(g('weight')),
            'market_value': safe_float(g('mktval')),
            'location':     g('location'),
            'currency':     g('currency'),
        })

    return holdings, scraped_date


def parse_vaneck_xlsx(filepath):
    """
    VanEck holdings XLSX format.
    Row 1 (index 0): 'All Holdings  MM/DD/YYYY'
    Row 2 (index 1): blank
    Row 3 (index 2): headers — Number, Holding Name, Ticker, ISIN, Shares, Market Value, % of Net Assets
    Row 4+ (index 3+): data
    Last row: 'Other/Cash' — skipped
    """
    holdings     = []
    scraped_date = None

    df = pd.read_excel(filepath, header=None)

    # Extract date from cell A1 e.g. "All Holdings  06/09/2026"
    cell_a1 = str(df.iloc[0, 0]) if not pd.isna(df.iloc[0, 0]) else ''
    m = re.search(r'(\d{2}/\d{2}/\d{4})', cell_a1)
    if m:
        try:
            scraped_date = datetime.strptime(m.group(1), '%m/%d/%Y').strftime('%Y-%m-%d')
        except:
            pass

    # Data rows start at index 3 (skip header row at index 2)
    data = df.iloc[3:].reset_index(drop=True)

    for _, row in data.iterrows():
        name       = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ''
        ticker     = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else ''
        mktval_raw = str(row.iloc[5]).strip() if not pd.isna(row.iloc[5]) else ''
        weight_raw = str(row.iloc[6]).strip() if not pd.isna(row.iloc[6]) else ''

        # Skip empty, cash, or placeholder rows
        if not name or name in ('nan', '--', 'Other/Cash', 'NaN'):
            continue
        if ticker.strip() in ('--', ''):
            ticker = ''

        weight = safe_float(weight_raw.replace('%', '').strip())
        mktval = safe_float(mktval_raw.replace('$', '').replace(',', '').strip())

        holdings.append({
            'ticker':       ticker,
            'name':         name,
            'sector':       None,
            'asset_class':  'Equity',
            'weight_pct':   weight,
            'market_value': mktval,
            'location':     None,
            'currency':     'USD',
        })

    return holdings, scraped_date



def parse_xtrackers_xlsx(filepath):
    """
    Xtrackers / DWS holdings XLSX format.
    Sheet name = date (e.g. '2026-06-18').
    Row 1: disclaimer text (skipped)
    Row 4 (index 3): headers — #, Name, ISIN, Country, Currency, Exchange,
                      Type of Security, Rating, Primary Listing,
                      Industry Classification, Weighting
    Row 5+ (index 4+): data
    No ticker column — ISIN only. Weighting is a decimal fraction (0.1164 = 11.64%).
    """
    holdings     = []
    scraped_date = None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = wb.sheetnames[0]

    # Sheet name is the as-of date, e.g. '2026-06-18'
    m = re.search(r'(\d{4}-\d{2}-\d{2})', sheet_name)
    if m:
        scraped_date = m.group(1)

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (contains 'Name' and 'ISIN')
    header_idx = None
    for i, row in enumerate(rows):
        if row and 'Name' in row and 'ISIN' in row:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 3  # fallback to known position

    headers = [str(h).strip() if h else '' for h in rows[header_idx]]
    col = {h: idx for idx, h in enumerate(headers)}

    for row in rows[header_idx + 1:]:
        if not row or row[col.get('Name', 1)] is None:
            continue
        name        = str(row[col['Name']]).strip()
        isin_val    = str(row[col['ISIN']]).strip() if col.get('ISIN') is not None and row[col['ISIN']] else ''
        currency    = str(row[col['Currency']]).strip() if col.get('Currency') is not None and row[col['Currency']] else None
        country     = str(row[col['Country']]).strip() if col.get('Country') is not None and row[col['Country']] else None
        sector      = str(row[col['Industry Classification']]).strip() if col.get('Industry Classification') is not None and row[col['Industry Classification']] else None
        sec_type    = str(row[col['Type of Security']]).strip() if col.get('Type of Security') is not None and row[col['Type of Security']] else ''
        weight_raw  = row[col['Weighting']] if col.get('Weighting') is not None else None

        if not name or name.lower() in ('nan', 'cash', 'other'):
            continue
        # Skip futures/derivatives/cash-management instruments
        if sec_type.lower() in ('future', 'futures', 'forward', 'swap', 'option', 'cash'):
            continue
        if isin_val.lower() in ('nan', '', 'none'):
            isin_val = ''

        try:
            weight = float(weight_raw) * 100 if weight_raw is not None else None
        except (TypeError, ValueError):
            weight = None

        holdings.append({
            'ticker':       '',           # Xtrackers provides no ticker — ISIN-based matching only
            'name':         name,
            'sector':       sector,
            'asset_class':  'Equity',
            'weight_pct':   weight,
            'market_value': None,
            'location':     country,
            'currency':     currency,
            'isin':         isin_val,
        })

    return holdings, scraped_date


def parse_globalx_csv(filepath):
    """
    Global X ETF holdings CSV — handles two formats:

    Format A (top-holdings): headers on row 1
      AS_OF_DATE, SEDOL, NAME, NET_ASSETS, TICKER, MARKET_PRICE, SHARES_HELD, MARKET_VALUE, ISIN, COUNTRY

    Format B (full-holdings): multi-line header
      Row 1: Fund name
      Row 2: "Fund Holdings Data as of MM/DD/YYYY"
      Row 3: % of Net Assets, Ticker, Name, SEDOL, Market Price, Shares Held, Market Value
    """
    holdings     = []
    scraped_date = None

    SKIP_NAMES = {'CASH', 'OTHER PAYABLE & RECEIVABLES', 'BRITISH STERLING POUND',
                  'EURO', 'HONG KONG DOLLAR', 'JAPANESE YEN', 'US DOLLAR',
                  'SINGAPORE DOLLAR', 'NORWEGIAN KRONE'}
    SKIP_PREFIXES = ('CURRENCY CONTRACT',)

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    if not lines:
        return [], None

    # Detect format by checking if first line looks like a header row
    first = lines[0].strip()
    if first.startswith('AS_OF_DATE'):
        # Format A — standard CSV with headers on row 1
        import io
        reader = csv.DictReader(io.StringIO(''.join(lines)))
        for row in reader:
            name     = str(row.get('NAME', '') or '').strip()
            sedol    = str(row.get('SEDOL', '') or '').strip()
            isin_val = str(row.get('ISIN', '') or '').strip()
            ticker   = str(row.get('TICKER', '') or '').strip()
            country  = str(row.get('COUNTRY', '') or '').strip()
            date_str = str(row.get('AS_OF_DATE', '') or '').strip()

            if not scraped_date and date_str:
                try:
                    scraped_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    pass

            if not name or name.upper() in SKIP_NAMES:
                continue
            if any(name.upper().startswith(p) for p in SKIP_PREFIXES):
                continue
            if not sedol and not isin_val:
                continue

            try:
                weight = float(row.get('NET_ASSETS', 0) or 0)
            except (ValueError, TypeError):
                weight = None

            holdings.append({
                'ticker': ticker, 'name': name, 'sector': None,
                'asset_class': 'Equity', 'weight_pct': weight,
                'market_value': None, 'location': country or None,
                'currency': None, 'isin': isin_val or None, 'sedol': sedol or None,
            })

    else:
        # Format B — multi-line header, date on row 2, headers on row 3
        # Extract date from row 2: "Fund Holdings Data as of MM/DD/YYYY"
        if len(lines) >= 2:
            date_line = lines[1].strip()
            m = re.search(r'(\d{2}/\d{2}/\d{4})', date_line)
            if m:
                try:
                    scraped_date = datetime.strptime(m.group(1), '%m/%d/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    pass

        # Headers on row 3 (index 2)
        import io
        reader = csv.DictReader(io.StringIO(''.join(lines[2:])))
        for row in reader:
            name   = str(row.get('Name', '') or '').strip()
            sedol  = str(row.get('SEDOL', '') or '').strip()
            ticker = str(row.get('Ticker', '') or '').strip()

            if not name or name.upper() in SKIP_NAMES:
                continue
            if any(name.upper().startswith(p) for p in SKIP_PREFIXES):
                continue
            if not sedol and not ticker:
                continue

            # Weight is "% of Net Assets" — may be negative for cash/FX
            weight_str = str(row.get('% of Net Assets', '') or '').replace(',', '').strip()
            try:
                weight = float(weight_str)
                if weight <= 0:
                    continue  # skip cash/FX/short positions
            except (ValueError, TypeError):
                continue

            sedol = sedol.strip('"').strip() if sedol else ''

            holdings.append({
                'ticker': ticker, 'name': name, 'sector': None,
                'asset_class': 'Equity', 'weight_pct': weight,
                'market_value': None, 'location': 'US',
                'currency': None, 'isin': None, 'sedol': sedol or None,
            })

    return holdings, scraped_date


def parse_hanetf_xlsx(filepath):
    """
    HANetf holdings XLSX format.
    Row 0: 'Fund Name As Of:DD-MM-YYYY'
    Row 3: headers — Security Description, Shares, Market Value (Base),
            Trading Currency, SEDOL/CUSIP, Exposure Country, Region, ISIN, Weight
    Row 4: blank
    Row 5+: data
    Weight is ratio — multiply by 100.
    """
    holdings     = []
    scraped_date = None

    df = pd.read_excel(filepath, header=None)

    # Extract date from cell A1 e.g. "Future of Defence UCITS ETF As Of:10-06-2026"
    cell_a1 = str(df.iloc[0, 0]) if not pd.isna(df.iloc[0, 0]) else ''
    m = re.search(r'(\d{2}-\d{2}-\d{4})', cell_a1)
    if m:
        try:
            scraped_date = datetime.strptime(m.group(1), '%d-%m-%Y').strftime('%Y-%m-%d')
        except:
            pass

    # Data starts at row 5, headers at row 3
    data = df.iloc[5:].reset_index(drop=True)

    for _, row in data.iterrows():
        name     = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
        mktval   = row.iloc[2]
        currency = str(row.iloc[3]).strip() if not pd.isna(row.iloc[3]) else ''
        sedol    = str(row.iloc[4]).strip() if not pd.isna(row.iloc[4]) else ''
        location = str(row.iloc[5]).strip() if not pd.isna(row.iloc[5]) else ''
        weight   = row.iloc[8]

        if not name or name in ('nan', 'NaN', ''):
            continue

        holdings.append({
            'ticker':       sedol,
            'name':         name,
            'sector':       None,
            'asset_class':  'Equity',
            'weight_pct':   safe_float(weight, multiply=100),
            'market_value': safe_float(mktval),
            'location':     location,
            'currency':     currency,
        })

    return holdings, scraped_date


def parse_wisdomtree_csv(filepath):
    """
    WisdomTree CSV format.
    Columns: Date, Fund Name, Fund Ticker Symbol, Security Description,
             Security Ticker, Shares, Mkt Value, Weight %
    Weight is ratio (0.057) — multiply by 100.
    Date is ISO format: 2026-06-09T00:00:00
    """
    holdings     = []
    scraped_date = None

    df = pd.read_csv(filepath)
    df.columns = [c.strip() for c in df.columns]  # strip leading spaces

    # Date from first row
    try:
        scraped_date = pd.to_datetime(df['Date'].iloc[0]).strftime('%Y-%m-%d')
    except:
        pass

    for _, row in df.iterrows():
        name   = str(row.get('Security Description', '')).strip()
        ticker = str(row.get('Security Ticker', '')).strip()

        if not name or name == 'nan':
            continue

        holdings.append({
            'ticker':       ticker,
            'name':         name,
            'sector':       None,
            'asset_class':  'Equity',
            'weight_pct':   safe_float(row.get('Weight %'), multiply=100),
            'market_value': safe_float(row.get('Mkt Value')),
            'location':     None,
            'currency':     'USD',
        })

    return holdings, scraped_date


# ── Historical Excel parsers ──────────────────────────────────────────────────

def parse_excel_funds_sheet(xl):
    """
    'Funds' sheet — iShares format with FUND column.
    Returns dict: {(etf_fund_id, date): [holdings]}
    """
    results = {}
    df      = xl.parse('Funds', header=0)
    df.columns = [str(c).strip() for c in df.columns]

    col_map = {
        df.columns[0]:  'FUND',
        df.columns[1]:  'Date',
        df.columns[2]:  'Ticker',
        df.columns[3]:  'Name',
        df.columns[5]:  'Sector',
        df.columns[6]:  'Asset Class',
        df.columns[7]:  'Market Value',
        df.columns[8]:  'Weight (%)',
        df.columns[11]: 'Location',
        df.columns[14]: 'Currency',
    }
    df = df.rename(columns=col_map)

    for _, row in df.iterrows():
        fund_code   = str(row.get('FUND', '')).strip().upper()
        etf_fund_id = FUND_ID_MAP.get(fund_code)
        if not etf_fund_id:
            continue

        try:
            scraped_date = pd.to_datetime(row.get('Date')).strftime('%Y-%m-%d')
        except:
            continue

        asset_cls = str(row.get('Asset Class', '')).strip()
        if asset_cls.lower() in SKIP_ASSET_CLASSES:
            continue

        name = str(row.get('Name', '')).strip()
        if not name or name == 'nan':
            continue

        key = (etf_fund_id, scraped_date)
        if key not in results:
            results[key] = []

        results[key].append({
            'ticker':       str(row.get('Ticker', '')).strip(),
            'name':         name,
            'sector':       str(row.get('Sector', '')).strip(),
            'asset_class':  asset_cls,
            'weight_pct':   safe_float(row.get('Weight (%)')),
            'market_value': safe_float(row.get('Market Value')),
            'location':     str(row.get('Location', '')).strip(),
            'currency':     str(row.get('Currency', '')).strip(),
        })

    return results


def parse_excel_wqtm_sheet(xl):
    """
    WisdomTree format: Date, Fund Name, Fund Ticker Symbol, Security Description,
    Security Ticker, Shares, Mkt Value, Weight %
    Weight is ratio — multiply by 100.
    """
    results     = {}
    etf_fund_id = FUND_ID_MAP.get('WQTM')
    if not etf_fund_id:
        return results

    df = xl.parse('WQTM', header=0)
    for _, row in df.iterrows():
        try:
            scraped_date = pd.to_datetime(row.iloc[0]).strftime('%Y-%m-%d')
        except:
            continue

        name = str(row.iloc[3]).strip()
        if not name or name == 'nan':
            continue

        key = (etf_fund_id, scraped_date)
        if key not in results:
            results[key] = []

        results[key].append({
            'ticker':       str(row.iloc[4]).strip(),
            'name':         name,
            'sector':       None,
            'asset_class':  'Equity',
            'weight_pct':   safe_float(row.iloc[7], multiply=100),
            'market_value': safe_float(row.iloc[6]),
            'location':     None,
            'currency':     None,
        })

    return results


def parse_excel_natp_sheet(xl):
    """
    NATP format: FUND, Date, Security Description, Shares, Market Value (Base),
    Trading Currency, SEDOL/CUSIP, Exposure Country, Region, ISIN, Weight
    Weight is ratio — multiply by 100.
    """
    results     = {}
    etf_fund_id = FUND_ID_MAP.get('NATP')
    if not etf_fund_id:
        return results

    df = xl.parse('NATP', header=0)
    df.columns = [str(c).strip() for c in df.columns]

    for _, row in df.iterrows():
        fund_code = str(row.iloc[0]).strip().upper()
        if fund_code != 'NATP':
            continue

        try:
            scraped_date = pd.to_datetime(row.iloc[1]).strftime('%Y-%m-%d')
        except:
            continue

        name = str(row.iloc[2]).strip()
        if not name or name == 'nan':
            continue

        key = (etf_fund_id, scraped_date)
        if key not in results:
            results[key] = []

        results[key].append({
            'ticker':       str(row.iloc[6]).strip(),
            'name':         name,
            'sector':       None,
            'asset_class':  'Equity',
            'weight_pct':   safe_float(row.iloc[10], multiply=100),
            'market_value': safe_float(row.iloc[4]),
            'location':     str(row.iloc[7]).strip(),
            'currency':     str(row.iloc[5]).strip(),
        })

    return results


def parse_excel_fcbr_sheet(xl):
    """
    FCBR format: ETF, Date, Security Name, Identifier, CUSIP, Classification,
    Shares or Quantity, Market Value, Weighting
    Weighting is ratio — multiply by 100.
    """
    results     = {}
    etf_fund_id = FUND_ID_MAP.get('FCBR')
    if not etf_fund_id:
        return results

    df = xl.parse('FCBR', header=0)
    for _, row in df.iterrows():
        fund_code = str(row.iloc[0]).strip().upper()
        if fund_code != 'FCBR':
            continue

        try:
            scraped_date = pd.to_datetime(row.iloc[1]).strftime('%Y-%m-%d')
        except:
            continue

        name = str(row.iloc[2]).strip()
        if not name or name == 'nan':
            continue

        key = (etf_fund_id, scraped_date)
        if key not in results:
            results[key] = []

        results[key].append({
            'ticker':       str(row.iloc[3]).strip(),
            'name':         name,
            'sector':       str(row.iloc[5]).strip(),
            'asset_class':  'Equity',
            'weight_pct':   safe_float(row.iloc[8], multiply=100),
            'market_value': safe_float(row.iloc[7]),
            'location':     None,
            'currency':     None,
        })

    return results


# ── Import logic ──────────────────────────────────────────────────────────────

def resolve_to_stock_map(conn):
    """
    For each new holding in etf_holdings, resolve to stock_identifier_map via OpenFIGI.
    - If already matched: skip
    - If found in OpenFIGI: insert with reviewed=1
    - If not found: insert UNRESOLVED placeholder with reviewed=0 for manual review
    """
    import requests, time

    FIGI_URL   = config.FIGI_URL
    FIGI_KEY   = config.OPENFIGI_API_KEY
    BATCH_SIZE = config.FIGI_BATCH_SIZE
    RATE_SLEEP = config.FIGI_RATE_SLEEP

    SUFFIX_MAP = {
        '.L':'LN', '.DE':'GY', '.PA':'FP', '.MI':'IM', '.AS':'NA',
        '.MC':'SM', '.ST':'SS', '.HE':'FH', '.KS':'KS', '.KQ':'KQ',
        '.T':'JT',  '.TW':'TT', '.HK':'HK', '.AX':'AU', '.OL':'NO',
        '.IS':'TI', '.E':'TI',  '.JK':'IJ', '.SZ':'CS', '.SS':'CH',
    }
    EXCH_TO_YAHOO = {
        'US':'', 'UQ':'', 'UN':'', 'UP':'', 'UA':'', 'UR':'',
        'LN':'.L', 'LX':'.L', 'GY':'.DE', 'GF':'.F', 'GS':'.SG',
        'FP':'.PA', 'IM':'.MI', 'NA':'.AS', 'SM':'.MC', 'SS':'.ST',
        'FH':'.HE', 'KS':'.KS', 'KQ':'.KQ', 'JT':'.T',  'TT':'.TW',
        'HK':'.HK', 'AU':'.AX', 'NO':'.OL', 'TI':'.IS', 'IJ':'.JK',
        'CS':'.SZ', 'CH':'.SS', 'DC':'.CO', 'BB':'.BR', 'PW':'.WA',
        'CN':'.TO', 'IT':'.TA', 'IN':'.NS', 'IB':'.BO', 'MK':'.KL',
    }
    PRIMARY = {'US','UQ','UN','UP','UA','LN','GY','FP','IM','NA',
               'SM','SS','FH','KS','KQ','JT','TT','HK','AU','NO',
               'TI','IJ','CS','CH','DC','BB','PW','CN','IT','IN','IB','MK'}

    # Location → preferred exchange code (from iShares Location column)
    LOCATION_TO_EXCH = {
        'taiwan':       'TT',
        'japan':        'JT',
        'china':        'HK',   # 4-digit numeric = HK-listed H-shares
        'hong kong':    'HK',
        'south korea':  'KS',
        'korea':        'KS',
        'germany':      'GY',
        'france':       'FP',
        'italy':        'IM',
        'netherlands':  'NA',
        'spain':        'SM',
        'sweden':       'SS',
        'finland':      'FH',
        'norway':       'NO',
        'turkey':       'TI',
        'indonesia':    'IJ',
        'australia':    'AU',
        'singapore':    'SP',
        'israel':       'IT',
        'india':        'IN',
        'malaysia':     'MK',
        'thailand':     'TB',
        'denmark':      'DC',
        'belgium':      'BB',
        'poland':       'PW',
        'canada':       'CN',
        'switzerland':  'SW',
        'brazil':       'BS',
        'mexico':       'MM',
        'south africa': 'SJ',
        'united kingdom': 'LN',
        'united states':  'US',
    }

    # Bloomberg exchange code → OpenFIGI exchange code
    # Covers codes used in WisdomTree/VanEck Bloomberg-format tickers
    BLOOMBERG_TO_OPENFIGI = {
        'US': 'US', 'UQ': 'UQ', 'UN': 'UN', 'UP': 'UP', 'UA': 'UA',
        'LN': 'LN', 'GY': 'GY', 'GR': 'GY', 'GF': 'GF', 'GS': 'GS',
        'FP': 'FP', 'IM': 'IM', 'NA': 'NA', 'SM': 'SM', 'SS': 'SS',
        'FH': 'FH', 'KS': 'KS', 'KQ': 'KQ', 'JT': 'JT', 'TT': 'TT',
        'HK': 'HK', 'AU': 'AU', 'NO': 'NO', 'TI': 'TI', 'IJ': 'IJ',
        'CS': 'CS', 'CH': 'CH', 'DC': 'DC', 'BB': 'BB', 'PW': 'PW',
        'CN': 'CN', 'CV': 'CV', 'IT': 'IT', 'IN': 'IN', 'IB': 'IB',
        'MK': 'MK', 'TB': 'TB', 'SP': 'SP', 'SJ': 'SJ',
        # Less common Bloomberg codes needing mapping
        'MT': 'NA',   # Malta → try Netherlands (ASML listed on Euronext Amsterdam)
        'SW': 'SW',   # Switzerland
        'VX': 'SW',   # SIX Swiss Exchange
        'SE': 'SS',   # Stockholm (alternate)
        'SF': 'FH',   # Helsinki (alternate)
        'FH': 'FH',
        'LI': 'LN',   # London International
        'E':  'TI',   # Borsa Istanbul old suffix (.E)
        'EI': 'ID',   # Ireland
        'ID': 'ID',
        'PL': 'PW',   # Warsaw (alternate)
        'BS': 'BS',   # Brazil
        'MM': 'MM',   # Mexico
        'CI': 'CI',   # Ivory Coast
        'NZ': 'NZ',   # New Zealand
        'JP': 'JT',   # Japan (alternate Bloomberg code)
    }

    def normalise_ticker(raw, location=None):
        """
        Returns list of OpenFIGI job dicts for this ticker — up to 3 attempts.
        Priority: location-based → suffix/bloomberg-code-based → bare ticker US
        """
        t    = str(raw).strip().upper()
        base = t.split()[0]
        attempts = []
        seen_exchcodes = set()

        def add_attempt(ticker_val, exch):
            if exch and exch not in seen_exchcodes:
                attempts.append({'idType': 'TICKER', 'idValue': ticker_val,
                                  'exchCode': exch, 'marketSecDes': 'Equity'})
                seen_exchcodes.add(exch)

        # Determine exchange from ticker suffix (e.g. '.L' → 'LN')
        exch_from_suffix = None
        clean_base       = base
        for suffix, code in SUFFIX_MAP.items():
            if clean_base.endswith(suffix.upper()):
                clean_base       = clean_base[:-len(suffix)]
                exch_from_suffix = code
                break

        # Handle Bloomberg space suffix: '028260 KS', 'ASELS TI', 'ASMLO MT'
        exch_from_bloomberg = None
        if ' ' in t:
            parts = t.split()
            if len(parts) == 2:
                candidate = BLOOMBERG_TO_OPENFIGI.get(parts[1])
                if candidate:
                    clean_base          = parts[0]
                    exch_from_bloomberg = candidate
                elif parts[1] in PRIMARY:
                    clean_base          = parts[0]
                    exch_from_bloomberg = parts[1]

        # Clean special chars for OpenFIGI
        clean_base = re.sub(r'-([A-Z])$', r'/\1', clean_base)  # MOG-A → MOG/A, SAAB-B → SAAB/B
        clean_base = re.sub(r'-', '/', clean_base)              # U-U → U/U (preserve as slash, not remove)
        if '.' in clean_base and not clean_base.endswith('/'):
            clean_base = clean_base.replace('.', '/')           # BA. → BA/

        # Attempt 1: location-based (most reliable for numeric tickers from iShares)
        if location:
            loc_exch = LOCATION_TO_EXCH.get(location.lower().strip())
            if loc_exch:
                add_attempt(clean_base, loc_exch)

        # Attempt 2: Bloomberg space code or suffix code
        if exch_from_bloomberg:
            add_attempt(clean_base, exch_from_bloomberg)
        elif exch_from_suffix:
            add_attempt(clean_base, exch_from_suffix)

        # Attempt 3: bare ticker with US as last resort
        add_attempt(clean_base, 'US')

        return attempts[:3]

    def pick_best(data):
        """Pick best equity result, preferring primary exchanges."""
        equities = [d for d in data
                   if d.get('marketSector') == 'Equity'
                   and d.get('securityType2') in ('Common Stock','ETP','ETF','Depositary Receipt')]
        primary  = [e for e in equities if e.get('exchCode') in PRIMARY]
        return primary[0] if primary else (equities[0] if equities else None)

    def insert_hit(conn, hit, original_ticker, original_name):
        figi        = hit['figi']
        fname       = hit.get('name', original_name)
        base_ticker = hit.get('ticker', '')
        exch_code   = hit.get('exchCode', '')
        bloomberg_c = f"{base_ticker} {exch_code}".strip()
        yahoo_id    = f"YF:{base_ticker}{EXCH_TO_YAHOO.get(exch_code, '')}"
        sec_type    = hit.get('securityType2', '')
        raw         = str(original_ticker).strip().upper()
        conn.execute("""
            INSERT INTO stock_identifier_map
                (figi, name, base_ticker, exch_code, bloomberg_code,
                 raw_ticker, yahoo_id, security_type, group_figi, reviewed)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ON CONFLICT(figi) DO UPDATE SET
                bloomberg_code = COALESCE(excluded.bloomberg_code, bloomberg_code),
                raw_ticker     = COALESCE(excluded.raw_ticker,     raw_ticker),
                yahoo_id       = COALESCE(excluded.yahoo_id,       yahoo_id),
                group_figi     = COALESCE(group_figi, excluded.group_figi)
        """, (figi, fname, base_ticker, exch_code, bloomberg_c,
              raw, yahoo_id, sec_type, figi))

    # Find all unresolved tickers with their location info
    all_tickers = conn.execute("""
        SELECT DISTINCT h.ticker, h.name, h.location
        FROM etf_holdings h
        WHERE h.ticker IS NOT NULL AND h.ticker != ''
    """).fetchall()

    unresolved = []
    for ticker, name, location in all_tickers:
        t  = str(ticker).strip().upper()
        tb = t.split()[0]
        exists = conn.execute("""
            SELECT 1 FROM stock_identifier_map
            WHERE figi NOT LIKE 'UNRESOLVED:%'
            AND (bloomberg_code = ? OR bloomberg_code = ?
            OR base_ticker = ? OR base_ticker = ?
            OR raw_ticker = ? OR raw_ticker = ?
            OR sedol = ? OR isin = ?)
            LIMIT 1
        """, (t, tb, t, tb, t, tb, t, t)).fetchone()
        if not exists:
            unresolved.append((ticker, name, location))

    if not unresolved:
        print("\n  All tickers already in stock_identifier_map")
        return

    print(f"\nResolving {len(unresolved)} new tickers via OpenFIGI...")
    inserted  = 0
    failed    = []
    headers   = {'Content-Type': 'application/json', 'X-OPENFIGI-APIKEY': FIGI_KEY}

    # Build flat job list with back-reference to original ticker
    # Each ticker can have up to 2 attempts
    job_map  = []  # list of (ticker, name, location, attempt_idx)
    job_list = []  # flat list of OpenFIGI jobs

    for ticker, name, location in unresolved:
        attempts = normalise_ticker(ticker, location)
        for idx, job in enumerate(attempts):
            job_list.append(job)
            job_map.append((ticker, name, location, idx))

    # Process in batches of BATCH_SIZE
    resolved_tickers = set()

    for i in range(0, len(job_list), BATCH_SIZE):
        batch_jobs = job_list[i:i+BATCH_SIZE]
        batch_meta = job_map[i:i+BATCH_SIZE]

        try:
            resp    = requests.post(FIGI_URL, headers=headers, json=batch_jobs, timeout=30)
            results = resp.json()
        except Exception as e:
            print(f"  OpenFIGI error: {e}")
            results = [None] * len(batch_jobs)

        for (ticker, name, location, attempt_idx), result in zip(batch_meta, results):
            if ticker in resolved_tickers:
                continue  # already resolved by a previous attempt

            hit = None
            if isinstance(result, dict) and result.get('data'):
                hit = pick_best(result['data'])

            if hit and hit.get('figi'):
                try:
                    insert_hit(conn, hit, ticker, name)
                    resolved_tickers.add(ticker)
                    inserted += 1
                except Exception as e:
                    print(f"  Insert error {ticker}: {e}")

        conn.commit()
        time.sleep(RATE_SLEEP)

    # Insert UNRESOLVED placeholders for anything still not matched
    for ticker, name, location in unresolved:
        if ticker not in resolved_tickers:
            failed.append(ticker)
            placeholder = f"UNRESOLVED:{ticker}|{name}"
            raw = str(ticker).strip().upper()
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO stock_identifier_map
                        (figi, name, bloomberg_code, raw_ticker, group_figi, reviewed)
                    VALUES (?, ?, ?, ?, ?, 0)
                """, (placeholder, name, raw, raw, placeholder))
            except:
                pass
    conn.commit()

    print(f"  Resolved: {inserted}, Unresolved: {len(failed)}")
    if failed:
        print(f"  First 10 unresolved: {failed[:10]}")

    # ── ISIN-only resolution (for providers like Xtrackers with no ticker column) ──
    isin_holdings = conn.execute("""
        SELECT DISTINCT h.isin, h.name
        FROM etf_holdings h
        WHERE (h.ticker IS NULL OR h.ticker = '')
        AND h.isin IS NOT NULL AND h.isin != ''
    """).fetchall()

    isin_unresolved = []
    for isin_val, name in isin_holdings:
        iv = str(isin_val).strip().upper()
        exists = conn.execute("""
            SELECT 1 FROM stock_identifier_map
            WHERE figi NOT LIKE 'UNRESOLVED:%' AND isin = ?
            LIMIT 1
        """, (iv,)).fetchone()
        if not exists:
            isin_unresolved.append((iv, name))

    if isin_unresolved:
        print(f"\nResolving {len(isin_unresolved)} new ISINs via OpenFIGI...")
        isin_resolved = 0
        isin_failed   = []

        isin_jobs = [{'idType': 'ID_ISIN', 'idValue': iv, 'marketSecDes': 'Equity'}
                     for iv, _ in isin_unresolved]

        for i in range(0, len(isin_jobs), BATCH_SIZE):
            batch_jobs = isin_jobs[i:i+BATCH_SIZE]
            batch_meta = isin_unresolved[i:i+BATCH_SIZE]
            try:
                resp    = requests.post(FIGI_URL, headers=headers, json=batch_jobs, timeout=30)
                results = resp.json()
            except Exception as e:
                print(f"  OpenFIGI error: {e}")
                results = [None] * len(batch_jobs)

            for (isin_val, name), result in zip(batch_meta, results):
                hit = None
                if isinstance(result, dict) and result.get('data'):
                    hit = pick_best(result['data'])
                if hit and hit.get('figi'):
                    try:
                        figi        = hit['figi']
                        fname       = hit.get('name') or name
                        bloomberg_c = f"{hit.get('ticker','')} {hit.get('exchCode','')}".strip()
                        sec_type    = hit.get('securityType') or hit.get('marketSector')
                        yahoo_id    = None
                        exch        = hit.get('exchCode')
                        if exch and hit.get('ticker'):
                            suf = EXCH_TO_YAHOO.get(exch, '')
                            yahoo_id = f"YF:{hit['ticker']}{suf}"
                        conn.execute("""
                            INSERT INTO stock_identifier_map
                                (figi, name, base_ticker, exch_code, bloomberg_code,
                                 isin, yahoo_id, security_type, group_figi, reviewed)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                            ON CONFLICT(figi) DO UPDATE SET
                                isin       = COALESCE(excluded.isin, isin),
                                yahoo_id   = COALESCE(excluded.yahoo_id, yahoo_id),
                                group_figi = COALESCE(group_figi, excluded.group_figi)
                        """, (figi, fname, hit.get('ticker'), exch, bloomberg_c,
                              isin_val, yahoo_id, sec_type, figi))
                        isin_resolved += 1
                    except Exception as e:
                        print(f"  Insert error {isin_val}: {e}")
                else:
                    isin_failed.append((isin_val, name))

            conn.commit()
            time.sleep(RATE_SLEEP)

        for isin_val, name in isin_failed:
            placeholder = f"UNRESOLVED:{isin_val}|{name}"
            try:
                conn.execute("""
                    INSERT OR IGNORE INTO stock_identifier_map
                        (figi, name, isin, group_figi, reviewed)
                    VALUES (?, ?, ?, ?, 0)
                """, (placeholder, name, isin_val, placeholder))
            except:
                pass
        conn.commit()

        print(f"  ISIN Resolved: {isin_resolved}, Unresolved: {len(isin_failed)}")
        if isin_failed:
            print(f"  First 10 unresolved ISINs: {[x[0] for x in isin_failed[:10]]}")


def process_holdings_dict(conn, results, source_label):
    total_inserted = 0
    total_skipped  = 0

    for (etf_fund_id, scraped_date), holdings in sorted(results.items()):
        if already_imported(conn, etf_fund_id, scraped_date):
            print(f"  SKIP {etf_fund_id} {scraped_date} — already imported")
            total_skipped += 1
            continue

        inserted, errors = insert_holdings(conn, etf_fund_id, scraped_date, holdings)
        total_inserted  += inserted
        status = f"{inserted} rows"
        if errors:
            status += f", {errors} errors"
        print(f"  {etf_fund_id} {scraped_date}: {status}")

    return total_inserted, total_skipped


def import_historical_excel(conn):
    if not EXCEL_PATH.exists():
        print(f"Excel file not found: {EXCEL_PATH}")
        return

    print(f"\nImporting historical data from {EXCEL_PATH}...")
    xl = pd.ExcelFile(EXCEL_PATH)
    total_inserted = 0

    print("\n  Sheet: Funds")
    results = parse_excel_funds_sheet(xl)
    inserted, skipped = process_holdings_dict(conn, results, 'Funds')
    total_inserted += inserted
    print(f"  → {inserted} rows inserted, {skipped} dates skipped")

    if 'WQTM' in xl.sheet_names:
        print("\n  Sheet: WQTM")
        results = parse_excel_wqtm_sheet(xl)
        inserted, skipped = process_holdings_dict(conn, results, 'WQTM')
        total_inserted += inserted
        print(f"  → {inserted} rows inserted, {skipped} dates skipped")

    if 'NATP' in xl.sheet_names:
        print("\n  Sheet: NATP")
        results = parse_excel_natp_sheet(xl)
        inserted, skipped = process_holdings_dict(conn, results, 'NATP')
        total_inserted += inserted
        print(f"  → {inserted} rows inserted, {skipped} dates skipped")

    if 'FCBR' in xl.sheet_names:
        print("\n  Sheet: FCBR")
        results = parse_excel_fcbr_sheet(xl)
        inserted, skipped = process_holdings_dict(conn, results, 'FCBR')
        total_inserted += inserted
        print(f"  → {inserted} rows inserted, {skipped} dates skipped")

    print(f"\nHistorical import complete: {total_inserted} total rows inserted")


def import_csv_files(conn):
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(
        list(IMPORT_DIR.glob('*.csv')) +
        list(IMPORT_DIR.glob('*.xls')) +
        list(IMPORT_DIR.glob('*.xlsx'))
    )

    if not files:
        print(f"\nNo files found in {IMPORT_DIR}")
        return

    print(f"\nProcessing {len(files)} files from {IMPORT_DIR}...")
    total_inserted = 0

    for filepath in files:
        prefix      = filepath.stem.split('_')[0].upper()
        etf_fund_id = FUND_ID_MAP.get(prefix)
        provider    = ETF_PROVIDER_MAP.get(prefix, 'ishares')

        if not etf_fund_id:
            print(f"\n  SKIP {filepath.name} — unknown prefix '{prefix}'")
            print(f"  Add an instrument row with this source_id and a provider, under Data → Instruments.")
            continue

        print(f"\n  {filepath.name} → {etf_fund_id} (provider: {provider})")

        try:
            ext = filepath.suffix.lower()
            if ext == '.csv' and provider == 'wisdomtree':
                holdings, scraped_date = parse_wisdomtree_csv(str(filepath))
            elif ext == '.csv' and provider == 'globalx':
                holdings, scraped_date = parse_globalx_csv(str(filepath))
            elif ext == '.csv':
                holdings, scraped_date = parse_ishares_csv(str(filepath))
            elif ext == '.xlsx' and provider == 'hanetf':
                holdings, scraped_date = parse_hanetf_xlsx(str(filepath))
            elif ext == '.xlsx' and provider == 'vaneck':
                holdings, scraped_date = parse_vaneck_xlsx(str(filepath))
            elif ext == '.xlsx' and provider == 'xtrackers':
                holdings, scraped_date = parse_xtrackers_xlsx(str(filepath))
            elif ext in ('.xls', '.xlsx'):
                holdings, scraped_date = parse_ishares_xls(str(filepath))
            else:
                print(f"  SKIP — unsupported extension: {ext}")
                continue
        except Exception as e:
            print(f"  ERROR parsing file: {e}")
            continue

        if not scraped_date:
            scraped_date = datetime.now().strftime('%Y-%m-%d')
            print(f"  WARNING: could not read date from file, using today: {scraped_date}")

        print(f"  Date: {scraped_date}, Holdings: {len(holdings)}")

        if already_imported(conn, etf_fund_id, scraped_date):
            print(f"  SKIP — {etf_fund_id} {scraped_date} already in DB")
            filepath.rename(ARCHIVE_DIR / filepath.name)
            continue

        inserted, errors = insert_holdings(conn, etf_fund_id, scraped_date, holdings)
        total_inserted  += inserted
        print(f"  Inserted: {inserted} rows" + (f", Errors: {errors}" if errors else ""))

        filepath.rename(ARCHIVE_DIR / filepath.name)
        print(f"  Archived to archive/{filepath.name}")

    print(f"\nCSV import complete: {total_inserted} total rows inserted")


def main():
    parser = argparse.ArgumentParser(description='ETF Holdings Importer')
    parser.add_argument('--historical', action='store_true',
                        help='Import from Funds Database.xlsx')
    args = parser.parse_args()

    conn = get_conn()
    create_tables(conn)

    if args.historical:
        import_historical_excel(conn)

    import_csv_files(conn)
    resolve_to_stock_map(conn)

    total = conn.execute("SELECT COUNT(*) FROM etf_holdings").fetchone()[0]
    funds = conn.execute("SELECT COUNT(DISTINCT etf_fund_id) FROM etf_holdings").fetchone()[0]
    dates = conn.execute("SELECT COUNT(DISTINCT scraped_date) FROM etf_holdings").fetchone()[0]
    print(f"\n=== DB Summary ===")
    print(f"  Total holdings rows: {total:,}")
    print(f"  ETFs tracked:        {funds}")
    print(f"  Unique dates:        {dates}")

    conn.close()


if __name__ == '__main__':
    main()