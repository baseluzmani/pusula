"""
ETF holdings file parsers, one per provider.

Each returns (holdings, scraped_date). holdings is a list of dicts with the
keys the etf_holdings table expects; scraped_date is read from inside the
file, never from the filename, because a file can be downloaded days after
its as-of date.

Lifted unchanged from scripts/import_etf_holdings.py. They work and are
fiddly in provider-specific ways, so this move is deliberately a move and not
a rewrite.
"""

from __future__ import annotations

import csv
import io
import re
import xml.etree.ElementTree as ET
from datetime import datetime

import openpyxl
import pandas as pd

# Asset classes that are not holdings.
SKIP_ASSET_CLASSES = {
    'cash and/or derivatives', 'cash', 'futures', 'options',
    'cash collateral and margins',
}

# Currencies, FX pairs and cash placeholders that appear in a ticker column.
SKIP_TICKERS = {
    'AUD', 'CAD', 'CHF', 'CZK', 'EUR', 'GBP', 'HKD', 'JPY', 'KRW',
    'MXN', 'MYR', 'PLN', 'SEK', 'SGD', 'TRY', 'TWD', 'USD', 'ZAR',
    'DKK', 'NOK', 'NZD', 'THB', 'IDR', 'INR', 'BRL',
    '-', '$', 'CASH', 'XONE',
}


def safe_float(val, multiply=1.0):
    if val is None:
        return None
    try:
        s = str(val).replace(',', '').replace('$', '').strip()
        if not s or s in ('-', 'N/A', 'nan', '--'):
            return None
        return float(s) * multiply
    except (TypeError, ValueError):
        return None


def parse_ishares_csv(filepath):
    """iShares CSV. Header rows at top carry the date; data follows the
    column header row."""
    holdings, scraped_date = [], None

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    for line in lines[:5]:
        m = re.search(r'(\d{2}/\w+/\d{4})', line)
        if m:
            try:
                scraped_date = datetime.strptime(
                    m.group(1), '%d/%b/%Y').strftime('%Y-%m-%d')
                break
            except ValueError:
                pass

    header_idx = None
    for i, line in enumerate(lines):
        if 'Ticker' in line and 'Name' in line and 'Weight' in line:
            header_idx = i
            break
    if header_idx is None:
        return holdings, scraped_date

    for row in csv.DictReader(lines[header_idx:]):
        name = (row.get('Name') or '').strip().strip('"')
        asset_cls = (row.get('Asset Class') or '').strip().strip('"')
        if not name or asset_cls.lower() in SKIP_ASSET_CLASSES:
            continue
        holdings.append({
            'ticker': (row.get('Ticker') or row.get('Issuer Ticker')
                       or '').strip().strip('"'),
            'name': name,
            'sector': (row.get('Sector') or '').strip().strip('"'),
            'asset_class': asset_cls,
            'weight_pct': safe_float(row.get('Weight (%)')),
            'market_value': safe_float(row.get('Market Value')),
            'location': (row.get('Location') or '').strip().strip('"'),
            'currency': (row.get('Market Currency') or '').strip().strip('"'),
        })
    return holdings, scraped_date


def parse_ishares_xls(filepath):
    """iShares .xls, which is really an XML Spreadsheet document."""
    holdings, scraped_date = [], None

    with open(filepath, 'rb') as f:
        content = f.read().decode('utf-8-sig')

    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
    root = ET.fromstring(content)
    rows = root.findall('.//ss:Row', ns)

    def get_vals(row):
        cells = row.findall('ss:Cell', ns)
        return [c.find('ss:Data', ns).text or ''
                if c.find('ss:Data', ns) is not None else '' for c in cells]

    for row in rows[:8]:
        for v in get_vals(row):
            m = re.search(r'(\d{2}[/-]\w+[/-]\d{4}|\d{4}-\d{2}-\d{2})',
                          str(v or ''))
            if m:
                for fmt in ('%d/%b/%Y', '%d-%b-%Y', '%Y-%m-%d'):
                    try:
                        scraped_date = datetime.strptime(
                            m.group(1), fmt).strftime('%Y-%m-%d')
                        break
                    except ValueError:
                        pass
            if scraped_date:
                break
        if scraped_date:
            break

    headers, header_idx = [], None
    for i, row in enumerate(rows):
        vals = get_vals(row)
        if any(v in ('Name', 'Ticker', 'Issuer Ticker') for v in vals):
            headers, header_idx = vals, i
            break
    if header_idx is None:
        return holdings, scraped_date

    def col(options):
        for n in options:
            if n in headers:
                return headers.index(n)
        return None

    idx = {'ticker': col(['Ticker', 'Issuer Ticker']), 'name': col(['Name']),
           'sector': col(['Sector']), 'asset': col(['Asset Class']),
           'weight': col(['Weight (%)']), 'mktval': col(['Market Value']),
           'location': col(['Location']), 'currency': col(['Market Currency'])}

    for row in rows[header_idx + 1:]:
        vals = get_vals(row)
        if not any(vals):
            continue

        def g(k):
            i = idx[k]
            return vals[i].strip() if i is not None and i < len(vals) else ''

        name, asset_cls = g('name'), g('asset')
        if not name or asset_cls.lower() in SKIP_ASSET_CLASSES:
            continue
        holdings.append({
            'ticker': g('ticker'), 'name': name, 'sector': g('sector'),
            'asset_class': asset_cls,
            'weight_pct': safe_float(g('weight')),
            'market_value': safe_float(g('mktval')),
            'location': g('location'), 'currency': g('currency'),
        })
    return holdings, scraped_date


def parse_vaneck_xlsx(filepath):
    """VanEck. A1 carries 'All Holdings  MM/DD/YYYY'; data from row 4."""
    holdings, scraped_date = [], None
    df = pd.read_excel(filepath, header=None)

    cell_a1 = str(df.iloc[0, 0]) if not pd.isna(df.iloc[0, 0]) else ''
    m = re.search(r'(\d{2}/\d{2}/\d{4})', cell_a1)
    if m:
        try:
            scraped_date = datetime.strptime(
                m.group(1), '%m/%d/%Y').strftime('%Y-%m-%d')
        except ValueError:
            pass

    for _, row in df.iloc[3:].reset_index(drop=True).iterrows():
        name = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ''
        ticker = str(row.iloc[2]).strip() if not pd.isna(row.iloc[2]) else ''
        mktval_raw = str(row.iloc[5]).strip() if not pd.isna(row.iloc[5]) else ''
        weight_raw = str(row.iloc[6]).strip() if not pd.isna(row.iloc[6]) else ''

        if not name or name in ('nan', '--', 'Other/Cash', 'NaN'):
            continue
        if ticker.strip() in ('--', ''):
            ticker = ''

        holdings.append({
            'ticker': ticker, 'name': name, 'sector': None,
            'asset_class': 'Equity',
            'weight_pct': safe_float(weight_raw.replace('%', '').strip()),
            'market_value': safe_float(
                mktval_raw.replace('$', '').replace(',', '').strip()),
            'location': None, 'currency': 'USD',
        })
    return holdings, scraped_date


def parse_xtrackers_xlsx(filepath):
    """Xtrackers / DWS. Sheet name is the as-of date. ISIN only, no ticker.
    Weighting is a fraction."""
    holdings, scraped_date = [], None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = wb.sheetnames[0]
    m = re.search(r'(\d{4}-\d{2}-\d{2})', sheet_name)
    if m:
        scraped_date = m.group(1)

    rows = list(wb[sheet_name].iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows):
        if row and 'Name' in row and 'ISIN' in row:
            header_idx = i
            break
    if header_idx is None:
        header_idx = 3

    headers = [str(h).strip() if h else '' for h in rows[header_idx]]
    col = {h: idx for idx, h in enumerate(headers)}

    def cell(row, key):
        i = col.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return None
        return str(row[i]).strip()

    for row in rows[header_idx + 1:]:
        if not row or row[col.get('Name', 1)] is None:
            continue
        name = str(row[col['Name']]).strip()
        sec_type = (cell(row, 'Type of Security') or '')
        if not name or name.lower() in ('nan', 'cash', 'other'):
            continue
        if sec_type.lower() in ('future', 'futures', 'forward', 'swap',
                                'option', 'cash'):
            continue

        isin_val = cell(row, 'ISIN') or ''
        if isin_val.lower() in ('nan', '', 'none'):
            isin_val = ''

        weight_raw = row[col['Weighting']] if col.get('Weighting') is not None \
            else None
        try:
            weight = float(weight_raw) * 100 if weight_raw is not None else None
        except (TypeError, ValueError):
            weight = None

        holdings.append({
            'ticker': '', 'name': name,
            'sector': cell(row, 'Industry Classification'),
            'asset_class': 'Equity', 'weight_pct': weight,
            'market_value': None, 'location': cell(row, 'Country'),
            'currency': cell(row, 'Currency'), 'isin': isin_val,
        })
    return holdings, scraped_date


def parse_globalx_csv(filepath):
    """Global X, two layouts: headers on row 1, or a multi-line header with
    the date on row 2."""
    holdings, scraped_date = [], None

    SKIP_NAMES = {'CASH', 'OTHER PAYABLE & RECEIVABLES',
                  'BRITISH STERLING POUND', 'EURO', 'HONG KONG DOLLAR',
                  'JAPANESE YEN', 'US DOLLAR', 'SINGAPORE DOLLAR',
                  'NORWEGIAN KRONE'}
    SKIP_PREFIXES = ('CURRENCY CONTRACT',)

    with open(filepath, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()
    if not lines:
        return [], None

    if lines[0].strip().startswith('AS_OF_DATE'):
        for row in csv.DictReader(io.StringIO(''.join(lines))):
            name = str(row.get('NAME', '') or '').strip()
            sedol = str(row.get('SEDOL', '') or '').strip()
            isin_val = str(row.get('ISIN', '') or '').strip()
            date_str = str(row.get('AS_OF_DATE', '') or '').strip()

            if not scraped_date and date_str:
                try:
                    scraped_date = datetime.strptime(
                        date_str, '%Y-%m-%d').strftime('%Y-%m-%d')
                except ValueError:
                    pass

            if not name or name.upper() in SKIP_NAMES:
                continue
            if any(name.upper().startswith(p) for p in SKIP_PREFIXES):
                continue
            if not sedol and not isin_val:
                continue

            try:
                weight = float(row.get('NET_ASSETS', 0) or 0)
            except (ValueError, TypeError):
                weight = None

            holdings.append({
                'ticker': str(row.get('TICKER', '') or '').strip(),
                'name': name, 'sector': None, 'asset_class': 'Equity',
                'weight_pct': weight, 'market_value': None,
                'location': str(row.get('COUNTRY', '') or '').strip() or None,
                'currency': None, 'isin': isin_val or None,
                'sedol': sedol or None,
            })
    else:
        if len(lines) >= 2:
            m = re.search(r'(\d{2}/\d{2}/\d{4})', lines[1].strip())
            if m:
                try:
                    scraped_date = datetime.strptime(
                        m.group(1), '%m/%d/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    pass

        for row in csv.DictReader(io.StringIO(''.join(lines[2:]))):
            name = str(row.get('Name', '') or '').strip()
            sedol = str(row.get('SEDOL', '') or '').strip()
            ticker = str(row.get('Ticker', '') or '').strip()

            if not name or name.upper() in SKIP_NAMES:
                continue
            if any(name.upper().startswith(p) for p in SKIP_PREFIXES):
                continue
            if not sedol and not ticker:
                continue

            weight_str = str(row.get('% of Net Assets', '') or '') \
                .replace(',', '').strip()
            try:
                weight = float(weight_str)
                if weight <= 0:
                    continue          # cash, FX and short lines
            except (ValueError, TypeError):
                continue

            holdings.append({
                'ticker': ticker, 'name': name, 'sector': None,
                'asset_class': 'Equity', 'weight_pct': weight,
                'market_value': None, 'location': 'US', 'currency': None,
                'isin': None, 'sedol': sedol.strip('"').strip() or None,
            })
    return holdings, scraped_date


def parse_hanetf_xlsx(filepath):
    """HANetf. A1 carries 'Fund Name As Of:DD-MM-YYYY'; data from row 6.
    Weight is a ratio."""
    holdings, scraped_date = [], None
    df = pd.read_excel(filepath, header=None)

    cell_a1 = str(df.iloc[0, 0]) if not pd.isna(df.iloc[0, 0]) else ''
    m = re.search(r'(\d{2}-\d{2}-\d{4})', cell_a1)
    if m:
        try:
            scraped_date = datetime.strptime(
                m.group(1), '%d-%m-%Y').strftime('%Y-%m-%d')
        except ValueError:
            pass

    for _, row in df.iloc[5:].reset_index(drop=True).iterrows():
        name = str(row.iloc[0]).strip() if not pd.isna(row.iloc[0]) else ''
        if not name or name in ('nan', 'NaN', ''):
            continue
        holdings.append({
            'ticker': str(row.iloc[4]).strip()
                      if not pd.isna(row.iloc[4]) else '',
            'name': name, 'sector': None, 'asset_class': 'Equity',
            'weight_pct': safe_float(row.iloc[8], multiply=100),
            'market_value': safe_float(row.iloc[2]),
            'location': str(row.iloc[5]).strip()
                        if not pd.isna(row.iloc[5]) else '',
            'currency': str(row.iloc[3]).strip()
                        if not pd.isna(row.iloc[3]) else '',
        })
    return holdings, scraped_date


def parse_wisdomtree_csv(filepath):
    """WisdomTree. Date in the first column, weight as a ratio."""
    holdings, scraped_date = [], None

    df = pd.read_csv(filepath)
    df.columns = [c.strip() for c in df.columns]

    try:
        scraped_date = pd.to_datetime(df['Date'].iloc[0]).strftime('%Y-%m-%d')
    except (KeyError, IndexError, ValueError):
        pass

    for _, row in df.iterrows():
        name = str(row.get('Security Description', '')).strip()
        if not name or name == 'nan':
            continue
        holdings.append({
            'ticker': str(row.get('Security Ticker', '')).strip(),
            'name': name, 'sector': None, 'asset_class': 'Equity',
            'weight_pct': safe_float(row.get('Weight %'), multiply=100),
            'market_value': safe_float(row.get('Mkt Value')),
            'location': None, 'currency': 'USD',
        })
    return holdings, scraped_date


# provider name -> {extension: parser}. import_file() falls back to the
# iShares parsers, which is what the provider column defaults to.
PARSERS = {
    'wisdomtree': {'.csv': parse_wisdomtree_csv},
    'globalx': {'.csv': parse_globalx_csv},
    'hanetf': {'.xlsx': parse_hanetf_xlsx},
    'vaneck': {'.xlsx': parse_vaneck_xlsx},
    'xtrackers': {'.xlsx': parse_xtrackers_xlsx},
}


def parser_for(provider: str, ext: str):
    """The parser for this provider and extension, or the iShares default."""
    ext = ext.lower()
    specific = PARSERS.get((provider or '').lower(), {}).get(ext)
    if specific:
        return specific
    if ext == '.csv':
        return parse_ishares_csv
    if ext in ('.xls', '.xlsx'):
        return parse_ishares_xls
    return None
